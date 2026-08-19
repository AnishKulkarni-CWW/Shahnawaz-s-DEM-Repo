"""
FIDELITAS — Report Export

Builds the exportable audit report in CSV and self-contained HTML form.
"""

import io
import re
import csv
from datetime import datetime
from .models import Status

STATUS_FILL_HEX = {
    Status.PASS: "C6F6D5",
    Status.FAIL: "FED7D7",
    Status.WARNING: "FEFCBF",
    Status.INFO: "BEE3F8",
    Status.MANUAL_REVIEW: "E9D8FD",
    Status.NOT_CHECKED: "E2E8F0",
}

# openpyxl raises IllegalCharacterError for raw control characters (ASCII
# 0-31 except tab/newline/carriage-return) in a cell value — these can show
# up in real audit data when messy source HTML/footer text is echoed back
# into a finding's title/actual/etc. Stripped here rather than letting one
# bad string crash the entire Excel export.
_ILLEGAL_XLSX_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _xlsx_safe(value):
    """Strip characters openpyxl refuses to write to a worksheet cell.
    Non-string values (numbers, None) pass through unchanged."""
    if isinstance(value, str):
        return _ILLEGAL_XLSX_CHARS_RE.sub("", value)
    return value

def issues_to_csv(issues: list) -> bytes:
    buf = io.StringIO()
    if not issues:
        return b""
    fieldnames = list(issues[0].to_row().keys())
    writer = csv.DictWriter(buf, fieldnames=fieldnames)
    writer.writeheader()
    for issue in issues:
        writer.writerow(issue.to_row())
    return buf.getvalue().encode("utf-8-sig")


def build_html_report(project_name: str, variant_name: str, scores: dict, issues: list, checklist_groups: list = None) -> str:
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    rows = "\n".join(
        f"<tr class='row-{i.status.value.lower()}'>"
        f"<td>{i.id}</td><td>{i.category}</td>"
        f"<td>{i.severity.code} — {i.severity.value}</td>"
        f"<td>STATUS: {i.status.label}</td>"
        f"<td>{_esc(i.title)}</td>"
        f"<td>{_esc(i.expected or '')}</td>"
        f"<td>{_esc(i.actual or '')}</td>"
        f"<td>{_esc(i.difference or '')}</td>"
        f"<td>{_esc(i.recommendation or '')}</td>"
        "</tr>"
        for i in issues
    )
    counts = scores["counts"]
    cat_rows = "\n".join(
        f"<tr><td>{cat}</td><td>{'N/A' if s is None else f'{s}%'}</td></tr>"
        for cat, s in scores["by_category"].items()
    )
    checklist_rows = "\n".join(
        f"<tr class='row-{g['status'].value.lower()}'>"
        f"<td>{_esc(g['item'].number)}</td>"
        f"<td>STATUS: {g['status'].label}</td>"
        f"<td>{_esc(g['item'].text_en)}</td>"
        f"<td>{_esc(g['item'].dev_status or '')}</td>"
        f"<td>{len(g['supporting'])}</td>"
        "</tr>"
        for g in (checklist_groups or [])
    )
    checklist_section = "" if not checklist_groups else f"""
<div class="card">
<h3>Checklist Items</h3>
<table>
<tr><th>#</th><th>Status</th><th>Checkpoint</th><th>Dev Status</th><th>Supporting Findings</th></tr>
{checklist_rows}
</table>
</div>
"""
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>FIDELITAS QA Report — {_esc(project_name)}</title>
<style>
body{{font-family:'Segoe UI',Arial,sans-serif;background:#0f1115;color:#e8e8ec;margin:0;padding:40px;}}
h1{{color:#fff;font-weight:700;letter-spacing:0.08em;text-transform:uppercase;font-size:22px;margin-bottom:4px;}}
h3{{letter-spacing:0.05em;text-transform:uppercase;font-size:13px;color:#9aa0ac;font-weight:700;}}
.masthead-rule{{border:none;border-top:1px solid #2a2d35;margin:10px 0 24px 0;}}
.sub{{color:#9aa0ac;margin-bottom:30px;font-size:12px;letter-spacing:0.03em;}}
.score{{font-size:56px;font-weight:700;color:#4ade80;}}
.chip{{display:inline-block;padding:3px 10px;border-radius:3px;font-size:11px;font-weight:700;letter-spacing:0.05em;margin-right:10px;font-family:'Courier New',monospace;}}
table{{border-collapse:collapse;width:100%;margin-top:20px;font-size:13px;}}
th,td{{border:1px solid #2a2d35;padding:8px 10px;text-align:left;vertical-align:top;}}
th{{background:#1a1d24;color:#c9cdd6;letter-spacing:0.04em;text-transform:uppercase;font-size:11px;}}
.row-fail{{background:#2a1418;}}
.row-warning{{background:#2a2414;}}
.row-pass{{background:#12211a;}}
.row-manual_review{{background:#1e1830;}}
.row-not_checked{{background:#181a1f;}}
.card{{background:#181b21;border:1px solid #2a2d35;border-radius:10px;padding:20px;margin-bottom:16px;}}
</style></head>
<body>
<h1>FIDELITAS</h1>
<hr class="masthead-rule">
<div class="sub">PROJECT: {_esc(project_name)} &nbsp;»&nbsp; VARIANT: {_esc(variant_name)} &nbsp;»&nbsp; GENERATED: {generated}</div>

<div class="card">
  <div class="score">{scores['overall'] if scores['overall'] is not None else 'N/A'}{'%' if scores['overall'] is not None else ''}</div>
  <div style="letter-spacing:0.04em;text-transform:uppercase;font-size:12px;color:#9aa0ac;">Overall QA Score</div>
  <p style="margin-top:16px;">
    <span class="chip" style="background:#{STATUS_FILL_HEX[Status.PASS]};color:#14532d;">PASS {counts.get(Status.PASS,0)}</span>
    <span class="chip" style="background:#{STATUS_FILL_HEX[Status.FAIL]};color:#7f1d1d;">FAIL {counts.get(Status.FAIL,0)}</span>
    <span class="chip" style="background:#{STATUS_FILL_HEX[Status.WARNING]};color:#713f12;">WARNING {counts.get(Status.WARNING,0)}</span>
    <span class="chip" style="background:#{STATUS_FILL_HEX[Status.MANUAL_REVIEW]};color:#581c87;">MANUAL REVIEW {counts.get(Status.MANUAL_REVIEW,0)}</span>
    <span class="chip" style="background:#{STATUS_FILL_HEX[Status.NOT_CHECKED]};color:#374151;">NOT CHECKED {counts.get(Status.NOT_CHECKED,0)}</span>
    &nbsp; TOTAL: {scores['total_checks']}
  </p>
</div>
{checklist_section}
<div class="card">
<h3>Category Scores</h3>
<table><tr><th>Category</th><th>Score</th></tr>{cat_rows}</table>
</div>

<div class="card">
<h3>All Findings</h3>
<table>
<tr><th>ID</th><th>Category</th><th>Severity</th><th>Status</th><th>Issue</th><th>Expected</th><th>Actual</th><th>Difference</th><th>Recommendation</th></tr>
{rows}
</table>
</div>

<p style="color:#666;font-size:11px;margin-top:30px;letter-spacing:0.02em;">GENERATED BY FIDELITAS &nbsp;»&nbsp; Items marked MANUAL REVIEW or NOT CHECKED were not auto-verified and require human confirmation — excluded from the score.</p>
</body></html>"""


def _esc(s: str) -> str:
    return (str(s) if s is not None else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def build_excel_report(project_name: str, variant_name: str, scores: dict, issues: list, checklist_groups: list = None) -> bytes:
    """Professional-formatted .xlsx: Summary sheet + Findings sheet, Arial
    throughout, status-colored fills, frozen header row, autofilter."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    bold = Font(name="Arial", bold=True, size=13)
    normal = Font(name="Arial", size=11)
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D3748")

    ws["A1"] = "FIDELITAS — Emailer QA Audit Report"
    ws["A1"].font = Font(name="Arial", bold=True, size=16)
    ws["A2"] = f"Project: {project_name or 'Unknown Project'}"
    ws["A2"].font = normal
    ws["A3"] = f"Variant: {variant_name}"
    ws["A3"].font = normal
    ws["A4"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A4"].font = normal
    ws["A6"] = "Overall QA Score"
    ws["A6"].font = bold
    ws["B6"] = _xlsx_safe(scores["overall"] if scores["overall"] is not None else "N/A")
    ws["B6"].font = Font(name="Arial", bold=True, size=16,
                          color="2F855A" if (scores["overall"] or 0) >= 90 else "C05621")

    ws["A8"] = "Category"
    ws["B8"] = "Score (%)"
    for c in ("A8", "B8"):
        ws[c].font = header_font
        ws[c].fill = header_fill
    row = 9
    for cat, sc in sorted(scores["by_category"].items(), key=lambda x: (x[1] is None, x[0])):
        ws.cell(row=row, column=1, value=_xlsx_safe(cat)).font = normal
        ws.cell(row=row, column=2, value=_xlsx_safe(sc if sc is not None else "Manual review only")).font = normal
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Status").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2, value="Count").font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    row += 1
    for status, count in scores["counts"].items():
        ws.cell(row=row, column=1, value=_xlsx_safe(f"STATUS: {status.label}")).font = normal
        ws.cell(row=row, column=2, value=_xlsx_safe(count)).font = normal
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    # ---- Checklist Items sheet (the checkpoints from the uploaded file,
    # positioned right after Summary so they're the first thing seen —
    # this is the primary structure of the audit, not an afterthought) ----
    if checklist_groups:
        ws_cl = wb.create_sheet("Checklist Items")
        cl_headers = ["#", "Status", "Checkpoint", "Checkpoint (JA)", "Dev Status", "Source Sheet", "Supporting Findings", "Note"]
        for col, h in enumerate(cl_headers, start=1):
            cell = ws_cl.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for r, g in enumerate(checklist_groups, start=2):
            item = g["item"]
            values = [_xlsx_safe(item.number), _xlsx_safe(f"STATUS: {g['status'].label}"), _xlsx_safe(item.text_en), _xlsx_safe(item.text_ja or ""),
                      _xlsx_safe(item.dev_status or ""), _xlsx_safe(item.source_sheet), len(g["supporting"]), _xlsx_safe(g["reason"] or "")]
            fill = PatternFill("solid", fgColor=STATUS_FILL_HEX.get(g["status"], "FFFFFF"))
            for c, val in enumerate(values, start=1):
                cell = ws_cl.cell(row=r, column=c, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        cl_widths = [5, 16, 46, 30, 12, 20, 10, 40]
        for i, w in enumerate(cl_widths, start=1):
            ws_cl.column_dimensions[get_column_letter(i)].width = w
        ws_cl.freeze_panes = "A2"
        ws_cl.auto_filter.ref = f"A1:{get_column_letter(len(cl_headers))}{len(checklist_groups) + 1}"

    # ---- Findings sheet ----
    ws2 = wb.create_sheet("Findings")
    headers = ["ID", "Category", "Severity", "Status", "Issue", "Expected", "Actual", "Difference", "Location", "Recommendation", "Workflow"]
    for col, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for r, issue in enumerate(issues, start=2):
        values = [issue.id, _xlsx_safe(issue.category), _xlsx_safe(f"{issue.severity.code} — {issue.severity.value}"),
                  _xlsx_safe(f"STATUS: {issue.status.label}"), _xlsx_safe(issue.title), _xlsx_safe(issue.expected or ""),
                  _xlsx_safe(issue.actual or ""), _xlsx_safe(issue.difference or ""), _xlsx_safe(issue.location or ""),
                  _xlsx_safe(issue.recommendation or ""), _xlsx_safe(issue.workflow_status)]
        fill = PatternFill("solid", fgColor=STATUS_FILL_HEX.get(issue.status, "FFFFFF"))
        for c, val in enumerate(values, start=1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [6, 14, 12, 12, 32, 28, 28, 28, 16, 32, 12]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(issues) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_report(project_name: str, variant_name: str, scores: dict, issues: list, checklist_groups: list = None) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleX", parent=styles["Title"], fontSize=20, spaceAfter=4)
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8, leading=10)
    normal = styles["Normal"]

    status_order = [Status.PASS, Status.FAIL, Status.WARNING, Status.MANUAL_REVIEW, Status.NOT_CHECKED]
    status_line = "  &nbsp;»&nbsp;  ".join(
        f'<font color="{s.color}"><b>{s.label}</b></font> {scores["counts"].get(s, 0)}' for s in status_order
    )

    story = [
        Paragraph("FIDELITAS", ParagraphStyle("Mast", parent=title_style, fontSize=22)),
        Paragraph("EMAILER QA AUDIT REPORT", ParagraphStyle("MastSub", parent=normal, fontSize=9,
                   textColor=colors.HexColor("#6B7280"), spaceAfter=8)),
        Paragraph(f"PROJECT: {project_name or 'Unknown Project'} &nbsp;»&nbsp; VARIANT: {variant_name} "
                   f"&nbsp;»&nbsp; GENERATED: {datetime.now().strftime('%Y-%m-%d %H:%M')}", normal),
        Spacer(1, 10),
        Paragraph(f"<b>OVERALL QA SCORE: {scores['overall'] if scores['overall'] is not None else 'N/A'}"
                   f"{'%' if scores['overall'] is not None else ''}</b>", styles["Heading2"]),
        Paragraph(status_line + f"  &nbsp;»&nbsp;  TOTAL: {scores['total_checks']}", normal),
        Spacer(1, 14),
    ]

    cat_data = [["Category", "Score"]] + [
        [cat, ("N/A" if sc is None else f"{sc}%")]
        for cat, sc in sorted(scores["by_category"].items(), key=lambda x: (x[1] is None, x[0]))
    ]
    cat_table = Table(cat_data, colWidths=[220, 100])
    cat_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D3748")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFC")]),
    ]))
    story += [cat_table, PageBreak()]

    if checklist_groups:
        story.append(Paragraph("Checklist Items", styles["Heading2"]))
        story.append(Spacer(1, 6))
        cg_header = ["#", "Status", "Checkpoint", "Dev Status", "Evidence"]
        cg_rows = [cg_header]
        cg_fill_map = {
            Status.PASS: colors.HexColor("#C6F6D5"), Status.FAIL: colors.HexColor("#FED7D7"),
            Status.WARNING: colors.HexColor("#FEFCBF"), Status.INFO: colors.HexColor("#BEE3F8"),
            Status.MANUAL_REVIEW: colors.HexColor("#E9D8FD"), Status.NOT_CHECKED: colors.HexColor("#E2E8F0"),
        }
        for g in checklist_groups:
            cg_rows.append([
                g["item"].number, g["status"].label,
                Paragraph(_pdf_esc(g["item"].text_en), small),
                g["item"].dev_status or "", str(len(g["supporting"])),
            ])
        cg_table = Table(cg_rows, colWidths=[24, 60, 300, 55, 45], repeatRows=1)
        cg_style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D3748")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E0")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
        for r, g in enumerate(checklist_groups, start=1):
            cg_style_cmds.append(("BACKGROUND", (1, r), (1, r), cg_fill_map.get(g["status"], colors.white)))
        cg_table.setStyle(TableStyle(cg_style_cmds))
        story += [cg_table, PageBreak()]

    story.append(Paragraph("All Findings", styles["Heading2"]))
    story.append(Spacer(1, 6))

    header = ["ID", "Category", "Sev", "Status", "Issue", "Expected", "Actual", "Difference", "Recommendation"]
    rows = [header]
    for i in issues:
        rows.append([
            str(i.id), i.category, i.severity.code, i.status.label,
            Paragraph(_pdf_esc(i.title), small),
            Paragraph(_pdf_esc(i.expected or ""), small),
            Paragraph(_pdf_esc(i.actual or ""), small),
            Paragraph(_pdf_esc(i.difference or ""), small),
            Paragraph(_pdf_esc(i.recommendation or ""), small),
        ])

    col_widths = [22, 55, 32, 42, 110, 100, 100, 100, 130]
    table = Table(rows, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D3748")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]
    fill_map_pdf = {
        Status.PASS: colors.HexColor("#C6F6D5"), Status.FAIL: colors.HexColor("#FED7D7"),
        Status.WARNING: colors.HexColor("#FEFCBF"), Status.INFO: colors.HexColor("#BEE3F8"),
        Status.MANUAL_REVIEW: colors.HexColor("#E9D8FD"), Status.NOT_CHECKED: colors.HexColor("#E2E8F0"),
    }  # pastel row-tint variants — deliberately lighter than Status.color (which is tuned for
       # small text/chip accents, not full-cell backgrounds in a dense table)
    for r, i in enumerate(issues, start=1):
        style_cmds.append(("BACKGROUND", (3, r), (3, r), fill_map_pdf.get(i.status, colors.white)))
    table.setStyle(TableStyle(style_cmds))
    story.append(table)

    doc.build(story)
    return buf.getvalue()


def _pdf_esc(s: str) -> str:
    return _esc(s)
