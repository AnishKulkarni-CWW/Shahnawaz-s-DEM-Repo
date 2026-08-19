"""
FIDELITAS — Checklist Parser

Parses the client checklist workbook (the "CheckList_..._DEM.xlsx" format).
This is NOT a generic Excel reader — it understands the specific layout
used in this checklist family:

  Main sheet
      Row containing "プロジェクト名" -> project name
      Row containing "Subject"        -> subject line per variant (cols C..G)
      Row containing "Directory（DEM)" -> directory URL per variant

  checklist sheet
      Bilingual (JP/EN) sign-off items, one per row, in a "Check item" block.

  Internal_checklist for * sheet
      Numbered technical QA items (Sr No, item text, Dev Status).

  DEM(n) / SC(n) sheets  <-- the actual source-of-truth content
      Sequential rows using marker cells in one column and the value in the
      next:
          "▼Mail Title："   (marker alone)  -> next non-empty row is the title
          "▼ALT："          + value on same row -> ALT text for one image block
          "▼CTA"            + value on same row -> CTA button text
          "▼URL"            + value on same row -> URL for the CTA immediately above

  images（AEM） sheet
      Image naming/dimension spec table (media, product, size code, filename,
      dimensions, target file size).

If the uploaded workbook does not match this layout, the parser degrades
gracefully: it still extracts whatever marker-based blocks it can find and
reports which sheets it could not interpret, rather than pretending the
whole file was understood.
"""

import re
from typing import Optional
import openpyxl

from .models import ReferenceModel, ReferenceVariant, ReferenceBlock, ChecklistItem


MARKER_ALT = "▼ALT"
MARKER_CTA = "▼CTA"
MARKER_URL = "▼URL"
MARKER_TITLE = "▼Mail Title"
MARKER_BLOCKTEXT = "▼BlockText"

# Strips a leading "▼ALT：" / "▼ALT:" style prefix when the marker and its
# value were entered in the SAME cell (observed in the "SC" sheet) rather
# than split across two columns like the DEM(n) sheets. Anchored to the
# start of the string so it can't accidentally eat "：" appearing later in
# genuine Japanese body text.
_COMBINED_ALT_RE = re.compile(r"^▼ALT[：:]\s*")
_TEMPLATE_TOKEN_RE = re.compile(r"^%%[\w]+%%$")


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _row_values(ws, row_idx, max_col=12):
    return [_clean(ws.cell(row=row_idx, column=c).value) for c in range(1, max_col + 1)]


class ChecklistParseWarning:
    def __init__(self, sheet, message):
        self.sheet = sheet
        self.message = message


def parse_checklist_workbook(path: str) -> tuple[ReferenceModel, list]:
    """Returns (ReferenceModel, list[ChecklistParseWarning])."""
    warnings: list = []
    wb = openpyxl.load_workbook(path, data_only=True)
    model = ReferenceModel()

    # ---- Main sheet: project name + subjects + directories ----
    if "Main" in wb.sheetnames:
        ws = wb["Main"]
        subjects, directories = [], []
        for r in range(1, ws.max_row + 1):
            vals = _row_values(ws, r)
            joined = " ".join(v for v in vals if v)
            if not model.project_name and "プロジェクト名" in joined:
                for v in vals[2:]:
                    if v:
                        model.project_name = v
                        break
            if joined.startswith("Subject") or joined.startswith("件名"):
                subjects = [v for v in vals[2:] if v]
            if "Directory" in (vals[0] or "") and "DEM" in (vals[0] or ""):
                # directories may be on this row or the very next row (base
                # paths without index.html) — prefer whichever row actually
                # contains http(s) URLs in the per-variant columns.
                this_row_urls = [v for v in vals[2:7] if v and v.startswith("http")]
                next_vals = _row_values(ws, r + 1)
                next_row_urls = [v for v in next_vals[2:7] if v and v.startswith("http")]
                if len(next_row_urls) >= len(this_row_urls):
                    directories = next_row_urls
                else:
                    directories = this_row_urls
        for i, subj in enumerate(subjects, start=1):
            model.variants.append(ReferenceVariant(
                name=f"DEM({i})",
                subject=subj,
                directory_url=directories[i - 1] if i - 1 < len(directories) else None,
            ))
    else:
        warnings.append(ChecklistParseWarning("Main", "Sheet not found — project metadata unavailable."))

    # ---- checklist sheet: sign-off items ----
    if "checklist" in wb.sheetnames:
        ws = wb["checklist"]
        for r in range(1, ws.max_row + 1):
            vals = _row_values(ws, r)
            label = vals[1] if len(vals) > 1 else None
            if label and "\n" in label and ("か" in label or "?" in label):
                ja, _, en = label.partition("\n")
                model.checklist_items.append(ChecklistItem(
                    number=str(len(model.checklist_items) + 1),
                    text_en=en.strip() or ja.strip(),
                    text_ja=ja.strip(),
                    source_sheet="checklist",
                ))
    else:
        warnings.append(ChecklistParseWarning("checklist", "Sign-off checklist sheet not found."))

    # ---- Internal_checklist sheet: technical dev QA items ----
    # Two discovery paths: (1) a sheet explicitly named "Internal_checklist
    # for *" (the original multi-sheet workbook format), and (2) a fallback
    # scan of every sheet not already claimed above for the same "Sr No /
    # item text / Dev Status" header shape under ANY sheet name — this is
    # what a standalone single-sheet checklist (e.g. "Sheet1") looks like,
    # and it's a completely valid, common way this checklist family shows
    # up on its own, without the full multi-sheet workbook around it.
    claimed_sheets = {"Main", "checklist"} | {s for s in wb.sheetnames if re.match(r"^(DEM|SC)(\(\d+\))?$", s)}
    internal_sheet_name = next((s for s in wb.sheetnames if s.lower().startswith("internal_checklist")), None)

    def _extract_checklist_items(ws, sheet_name: str) -> int:
        header_row = None
        for r in range(1, min(ws.max_row, 5) + 1):
            vals = _row_values(ws, r)
            if vals and vals[0] and "sr no" in str(vals[0]).lower():
                header_row = r
                break
        if header_row is None:
            return 0
        added = 0
        start = header_row + 1
        for r in range(start, ws.max_row + 1):
            vals = _row_values(ws, r, max_col=5)
            if not vals[0] and not vals[1]:
                continue
            if vals[1]:
                model.checklist_items.append(ChecklistItem(
                    number=str(vals[0] or len(model.checklist_items) + 1),
                    text_en=vals[1],
                    dev_status=vals[2],
                    source_sheet=sheet_name,
                ))
                added += 1
        return added

    if internal_sheet_name:
        _extract_checklist_items(wb[internal_sheet_name], internal_sheet_name)
        claimed_sheets.add(internal_sheet_name)
    else:
        warnings.append(ChecklistParseWarning("Internal_checklist", "No sheet named 'Internal_checklist for *' found — scanning all sheets for a matching checklist layout instead."))

    for sheet_name in wb.sheetnames:
        if sheet_name in claimed_sheets or "image" in sheet_name.lower():
            continue
        added = _extract_checklist_items(wb[sheet_name], sheet_name)
        if added:
            claimed_sheets.add(sheet_name)

    # ---- DEM(n) / SC sheets: sequential source-of-truth content blocks ----
    # NOTE: "SC" appears both as bare "SC" (a full parallel variant, with its
    # own directory in the Main sheet) and would appear as "SC(n)" in other
    # checklist families — match both, don't require the numbered suffix.
    dem_sheets = [s for s in wb.sheetnames if re.match(r"^(DEM|SC)(\(\d+\))?$", s)]
    if not dem_sheets:
        warnings.append(ChecklistParseWarning("DEM(n)", "No DEM(n)/SC(n) content sheets found — cannot build reference content model."))

    for sheet_name in dem_sheets:
        ws = wb[sheet_name]
        variant = next((v for v in model.variants if v.name == sheet_name), None)
        if variant is None:
            variant = ReferenceVariant(name=sheet_name)
            model.variants.append(variant)

        order = 0
        pending_cta_text = None
        r = 1
        expecting_title_next = False
        while r <= ws.max_row:
            vals = _row_values(ws, r, max_col=6)
            # columns are 0-indexed here: D=index3, E=index4
            col_d = vals[3] if len(vals) > 3 else None
            col_e = vals[4] if len(vals) > 4 else None

            if col_e and MARKER_TITLE in col_e:
                expecting_title_next = True
                r += 1
                continue

            if expecting_title_next:
                if col_e:
                    order += 1
                    variant.blocks.append(ReferenceBlock(order=order, kind="title", text=col_e))
                    expecting_title_next = False
                r += 1
                continue

            if col_d and MARKER_ALT in col_d and col_e:
                order += 1
                variant.blocks.append(ReferenceBlock(order=order, kind="alt", text=col_e))

            elif col_d and MARKER_ALT in col_d and not col_e:
                # combined marker+value in one cell, e.g. "▼ALT：ニューBMW2..."
                extracted = _COMBINED_ALT_RE.sub("", col_d).strip()
                if extracted:
                    order += 1
                    variant.blocks.append(ReferenceBlock(order=order, kind="alt", text=extracted))

            elif col_d and MARKER_BLOCKTEXT in col_d and col_e:
                # Dynamic/templated content placeholder (e.g. "%%block1%%") —
                # NOT literal text. Recorded as its own kind so downstream
                # engines never try to string-match it against real HTML,
                # which would guarantee a false failure on every run.
                order += 1
                variant.blocks.append(ReferenceBlock(order=order, kind="blocktext", text=col_e))
                # a block reference is sometimes immediately followed by a
                # continuation row: no marker repeated, just another bare
                # "%%token%%" value in the same column.
                r += 1
                while r <= ws.max_row:
                    cont_vals = _row_values(ws, r, max_col=6)
                    cont_d = cont_vals[3] if len(cont_vals) > 3 else None
                    cont_e = cont_vals[4] if len(cont_vals) > 4 else None
                    if not cont_d and cont_e and _TEMPLATE_TOKEN_RE.match(cont_e.strip()):
                        order += 1
                        variant.blocks.append(ReferenceBlock(order=order, kind="blocktext", text=cont_e))
                        r += 1
                    else:
                        break
                continue

            elif col_d and MARKER_CTA in col_d and col_e:
                pending_cta_text = col_e

            elif col_d and MARKER_URL in col_d and col_e:
                order += 1
                variant.blocks.append(ReferenceBlock(
                    order=order, kind="cta",
                    text=pending_cta_text or "(CTA text not found)",
                    url=col_e,
                ))
                pending_cta_text = None

            r += 1

        if not variant.blocks:
            warnings.append(ChecklistParseWarning(sheet_name, "No ALT/CTA/Title marker blocks detected on this sheet."))

    # ---- images（AEM） sheet: naming/dimension spec ----
    # NOTE: parenthesized explicitly — this used to rely on Python's "and"
    # binding tighter than "or" to accidentally exclude the "Image size"
    # legend sheet from matching here. That was correct by luck, not by
    # design, so it's written unambiguously now.
    img_sheet_name = next(
        (s for s in wb.sheetnames if "images" in s.lower() or ("image" in s.lower() and "size" not in s.lower())),
        None,
    )
    if img_sheet_name and img_sheet_name in wb.sheetnames:
        ws = wb[img_sheet_name]
        header_row = None
        for r in range(1, min(ws.max_row, 8) + 1):
            vals = _row_values(ws, r, max_col=10)
            if vals and any(v and v.lower() == "media" for v in vals if v):
                header_row = r
                break
        if header_row:
            for r in range(header_row + 1, ws.max_row + 1):
                vals = _row_values(ws, r, max_col=10)
                # media, image, product, #, size, filename, notes, data-size
                # ("notes" frequently carries the actual WxH pixel dimension
                # embedded within free Japanese text, e.g. "フォームページ
                # TOPにも使用\n1680x756" — image_engine extracts it with a
                # regex search rather than assuming a clean bare value).
                if len(vals) > 6 and vals[6]:
                    model.image_specs.append({
                        "media": vals[1],
                        "product": vals[3],
                        "code": vals[5],
                        "filename": vals[6],
                        "dimensions": vals[7],
                        "target_size": vals[8] if len(vals) > 8 else None,
                    })

    # ---- Image size sheet: size-code -> pixel-dimension legend ----
    # Used two ways downstream: (1) as a fallback when a spec row's "notes"
    # field has no extractable WxH pattern, and (2) as a cross-check — if a
    # row's code says "xl" but its notes-embedded dimension disagrees with
    # what the legend defines for "xl", that is itself a real finding (the
    # checklist's own internal consistency, independent of the HTML).
    if "Image size" in wb.sheetnames:
        ws = wb["Image size"]
        for r in range(1, ws.max_row + 1):
            vals = _row_values(ws, r, max_col=6)
            code = vals[1] if len(vals) > 1 else None
            dims = vals[2] if len(vals) > 2 else None
            if code and dims and re.search(r"\d+\s*[x×]\s*\d+", str(dims)):
                model.image_size_legend[str(code).strip().lower()] = str(dims).strip()

    return model, warnings
